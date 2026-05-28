"""导出服务"""

import json
import logging
from io import BytesIO
from pathlib import Path

from fpdf import FPDF
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

from ..utils.error_handler import ErrorCategory, log_error

logger = logging.getLogger(__name__)

# 中文字体路径（项目内置SimHei黑体，回退系统字体）
FONT_DIR = Path(__file__).parent.parent.parent / 'static' / 'fonts'
CHINESE_FONT = FONT_DIR / 'simhei.ttf'
_SYSTEM_FONT_CANDIDATES = [
    Path('C:/Windows/Fonts/simhei.ttf'),
    Path('C:/Windows/Fonts/msyh.ttc'),
    Path('/usr/share/fonts/truetype/noto/NotoSansCJK-Regular.ttc'),
    Path('/System/Library/Fonts/PingFang.ttc'),
]


class ExportService:
    """导出服务类"""

    def _init_pdf_font(self, pdf: FPDF) -> bool:
        """初始化PDF中文字体（项目字体 -> 系统字体 -> 回退ASCII）"""
        # 1. 尝试项目内置字体
        if CHINESE_FONT.exists():
            try:
                pdf.add_font('SimHei', '', str(CHINESE_FONT))
                pdf.add_font('SimHei', 'B', str(CHINESE_FONT))
                return True
            except Exception as e:
                log_error(ErrorCategory.UNKNOWN, f'加载项目中文字体失败: {e}', level='warning')
        # 2. 尝试系统字体
        for font_path in _SYSTEM_FONT_CANDIDATES:
            if font_path.exists():
                try:
                    pdf.add_font('SimHei', '', str(font_path))
                    pdf.add_font('SimHei', 'B', str(font_path))
                    logger.info(f'使用系统中文字体: {font_path}')
                    return True
                except Exception:
                    continue
        logger.warning('未找到可用的中文字体，PDF将仅支持ASCII字符')
        return False
        logger.warning(f'中文字体文件不存在: {CHINESE_FONT}')
        return False

    def export_weekly_report_pdf(self, report) -> BytesIO | None:
        """导出周报为PDF

        Args:
            report: 周报对象

        Returns:
            BytesIO: PDF文件流
        """
        try:
            # 创建PDF对象
            pdf = FPDF()
            pdf.set_auto_page_break(auto=True, margin=15)

            # 添加中文字体支持
            has_chinese_font = self._init_pdf_font(pdf)
            font_name = 'SimHei' if has_chinese_font else 'Arial'

            pdf.add_page()

            # 标题
            pdf.set_font(font_name, '', 16)
            pdf.cell(0, 10, report.title, new_x='LMARGIN', new_y='NEXT', align='C')

            # 元数据
            pdf.set_font(font_name, '', 10)
            pdf.cell(
                0,
                8,
                f'发布日期: {report.report_date.strftime("%Y年%m月%d日")}',
                new_x='LMARGIN',
                new_y='NEXT',
                align='C',
            )
            pdf.cell(
                0,
                8,
                f'统计周期: {report.week_start.strftime("%Y-%m-%d")} 至 {report.week_end.strftime("%Y-%m-%d")}',
                new_x='LMARGIN',
                new_y='NEXT',
                align='C',
            )
            pdf.ln(10)

            # 摘要
            pdf.set_font(font_name, 'B', 12)
            pdf.cell(0, 10, '本周概览', new_x='LMARGIN', new_y='NEXT', align='L')
            pdf.set_font(font_name, '', 10)
            pdf.multi_cell(0, 5, report.summary)
            pdf.ln(10)

            # 详细内容
            if report.content:
                content = json.loads(report.content)

                # 重要变化
                if content.get('top_changes'):
                    pdf.set_font(font_name, 'B', 12)
                    pdf.cell(0, 10, '重要变化', new_x='LMARGIN', new_y='NEXT', align='L')
                    pdf.set_font(font_name, '', 10)
                    for change in content['top_changes']:
                        pdf.cell(
                            0, 6, f'• {change["title"]} - {change["author"]}', new_x='LMARGIN', new_y='NEXT', align='L'
                        )
                        pdf.cell(0, 6, f'  类别: {change["category"]}', new_x='LMARGIN', new_y='NEXT', align='L')
                        if change['rank_change'] > 0:
                            pdf.cell(
                                0,
                                6,
                                f'  排名变化: ↑ {change["rank_change"]} 位',
                                new_x='LMARGIN',
                                new_y='NEXT',
                                align='L',
                            )
                        elif change['rank_change'] < 0:
                            pdf.cell(
                                0,
                                6,
                                f'  排名变化: ↓ {abs(change["rank_change"])} 位',
                                new_x='LMARGIN',
                                new_y='NEXT',
                                align='L',
                            )
                        else:
                            pdf.cell(0, 6, '  排名变化: → 无变化', new_x='LMARGIN', new_y='NEXT', align='L')
                        pdf.ln(2)
                    pdf.ln(5)

                # 推荐书籍
                if content.get('featured_books'):
                    pdf.set_font(font_name, 'B', 12)
                    pdf.cell(0, 10, '推荐书籍', new_x='LMARGIN', new_y='NEXT', align='L')
                    pdf.set_font(font_name, '', 10)
                    for book in content['featured_books']:
                        pdf.cell(
                            0, 6, f'• {book["title"]} - {book["author"]}', new_x='LMARGIN', new_y='NEXT', align='L'
                        )
                        pdf.cell(0, 6, f'  推荐理由: {book["reason"]}', new_x='LMARGIN', new_y='NEXT', align='L')
                        pdf.ln(2)
                    pdf.ln(5)

            # 页脚
            pdf.set_font(font_name, '', 8)
            pdf.cell(
                0,
                10,
                f'© {report.report_date.year} BookRank - 纽约时报畅销书排行榜',
                new_x='LMARGIN',
                new_y='NEXT',
                align='C',
            )

            # 输出到内存流
            buffer = BytesIO()
            pdf.output(buffer)
            buffer.seek(0)

            logger.info(f'PDF导出成功: {report.title}')
            return buffer

        except Exception as e:
            log_error(ErrorCategory.UNKNOWN, f'PDF导出失败: {e!s}')
            return None

    def export_weekly_report_excel(self, report) -> BytesIO | None:
        """导出周报为Excel

        Args:
            report: 周报对象

        Returns:
            BytesIO: Excel文件流
        """
        try:
            # 创建工作簿
            wb = Workbook()
            ws = wb.active
            ws.title = '周报'

            # 设置列宽
            ws.column_dimensions['A'].width = 50
            ws.column_dimensions['B'].width = 30
            ws.column_dimensions['C'].width = 20
            ws.column_dimensions['D'].width = 15

            # 标题
            title_font = Font(bold=True, size=14)
            ws.merge_cells('A1:D1')
            ws['A1'] = report.title
            ws['A1'].font = title_font
            ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

            # 元数据
            meta_font = Font(size=10)
            ws['A3'] = f'发布日期: {report.report_date.strftime("%Y年%m月%d日")}'
            ws['A4'] = f'统计周期: {report.week_start.strftime("%Y-%m-%d")} 至 {report.week_end.strftime("%Y-%m-%d")}'
            ws['A3'].font = meta_font
            ws['A4'].font = meta_font

            # 摘要
            summary_font = Font(bold=True, size=12)
            ws['A6'] = '本周概览'
            ws['A6'].font = summary_font
            ws.merge_cells('A7:D7')
            ws['A7'] = report.summary
            ws['A7'].alignment = Alignment(wrap_text=True, vertical='top')

            # 详细内容
            if report.content:
                content = json.loads(report.content)
                row = 9

                # 重要变化
                if content.get('top_changes'):
                    ws[f'A{row}'] = '重要变化'
                    ws[f'A{row}'].font = summary_font
                    row += 1

                    # 表头
                    header_font = Font(bold=True)
                    ws[f'A{row}'] = '书名'
                    ws[f'B{row}'] = '作者'
                    ws[f'C{row}'] = '类别'
                    ws[f'D{row}'] = '排名变化'
                    for col in ['A', 'B', 'C', 'D']:
                        ws[f'{col}{row}'].font = header_font
                        ws[f'{col}{row}'].alignment = Alignment(horizontal='center')
                    row += 1

                    # 数据
                    for change in content['top_changes']:
                        ws[f'A{row}'] = change['title']
                        ws[f'B{row}'] = change['author']
                        ws[f'C{row}'] = change['category']
                        if change['rank_change'] > 0:
                            ws[f'D{row}'] = f'↑ {change["rank_change"]}'
                        elif change['rank_change'] < 0:
                            ws[f'D{row}'] = f'↓ {abs(change["rank_change"])}'
                        else:
                            ws[f'D{row}'] = '→ 无变化'
                        row += 1
                    row += 1

                # 推荐书籍
                if content.get('featured_books'):
                    ws[f'A{row}'] = '推荐书籍'
                    ws[f'A{row}'].font = summary_font
                    row += 1

                    # 表头
                    ws[f'A{row}'] = '书名'
                    ws[f'B{row}'] = '作者'
                    ws[f'C{row}'] = '推荐理由'
                    for col in ['A', 'B', 'C']:
                        ws[f'{col}{row}'].font = header_font
                        ws[f'{col}{row}'].alignment = Alignment(horizontal='center')
                    row += 1

                    # 数据
                    for book in content['featured_books']:
                        ws[f'A{row}'] = book['title']
                        ws[f'B{row}'] = book['author']
                        ws[f'C{row}'] = book['reason']
                        ws[f'C{row}'].alignment = Alignment(wrap_text=True, vertical='top')
                        row += 1

            # 页脚
            footer_font = Font(size=8)
            ws[f'A{row + 2}'] = f'© {report.report_date.year} BookRank - 纽约时报畅销书排行榜'
            ws[f'A{row + 2}'].font = footer_font
            ws.merge_cells(f'A{row + 2}:D{row + 2}')
            ws[f'A{row + 2}'].alignment = Alignment(horizontal='center')

            # 输出到内存流
            buffer = BytesIO()
            wb.save(buffer)
            buffer.seek(0)

            logger.info(f'Excel导出成功: {report.title}')
            return buffer

        except Exception as e:
            log_error(ErrorCategory.UNKNOWN, f'Excel导出失败: {e!s}')
            return None
